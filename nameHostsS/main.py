import openpyxl as xl

doc = xl.load_workbook('./exEmails.xlsx')

ws = doc.active

# for x in range(1,101):
#     for y in range(1,101):
#         print(ws.cell(row=x, column=y))
d = ws.cell(row=2, column=2, value=10)
print(d)